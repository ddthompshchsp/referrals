# referrals_app.py
from datetime import datetime, date
from zoneinfo import ZoneInfo
import re

import pandas as pd
import streamlit as st
from PIL import Image

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ==============================
# APP SETUP
# ==============================
st.set_page_config(page_title="Referrals Formatter", layout="centered")

try:
    logo = Image.open("header_logo.png")
    st.image(logo, width=300)
except Exception:
    pass

st.title("HCHSP Referrals Formatter (2025–2026)")
st.markdown("Upload your **Referrals/Service Tracking** workbook to receive a formatted report with charts.")

# ==============================
# EDITABLE MAPPINGS
# ==============================
# Map many Detailed Services into a single General Service bucket
GENERAL_DETAIL_MAP = {
    "Family Services - Emergency/Crisis": [
        "Utilities (No water/electricity)", "Fire", "Flood", "Homeless", "No Food", "No Clothing",
        "Emergency", "Crisis",
    ],
    "Family Well-Being (FPA)": [
        "Assistance Paying for Utilities", "Postpartum", "Employment Assistance", "Home Repairs",
        "Housing Assistance", "Reproductive Health", "SNAP", "TANF", "SSI", "Substance Abuse", "Tobacco",
    ],
    "Families as Lifelong Educators (FPA)": [
        "FSW Home Visit", "School Readiness", "Volunteer", "Parent Involvement",
    ],
    "Families as Learners (FPA)": [
        "Adult Education", "Job Training", "College", "Literacy", "Career Counseling",
        "ESL", "Financial Assistance for Education", "Scholarship", "GED", "Abriendo Puertas", "Parent Meeting",
    ],
    "Positive Parent-Child Relationship (FPA)": [
        "Parenting Education", "Parenting Skills", "Research-Based Parenting",
    ],
    "Family Engagement in Transition (FPA)": [
        "Early Childhood Intervention", "ECI", "Transition", "Supporting Transition",
    ],
    "Family Connections to Peers and Community (FPA)": [
        "Incarcerated", "Order of Protection", "Child Care", "CPS", "Child Support",
        "Marriage", "Family Relationships", "Foster Care", "Adoption", "Immigration",
        "Separation", "Divorce", "Legal", "Social Support",
    ],
    "Families as Advocates and Leaders (FPA)": [
        "Community Involvement", "Parent Committee", "Policy Council",
    ],
    "Education": [
        "Parent Teacher Conference", "PTC", "Education Service",
    ],
    "Disability": [
        "Suspected IEP", "IFP", "Evaluation", "Disability",
    ],
    "Health": [
        "Medical Condition", "Family Health Education", "Family Dental Education",
        "Insurance", "Blood Pressure", "Dental", "Dental Home", "Hearing", "Height", "Weight",
        "Lead", "Medical Home", "Physical", "TB", "Vision", "Immunization",
    ],
    "Mental Health": [
        "Behavior", "Counseling with HCHSP Staff", "Mental Health Education",
        "Rapid Response", "Consulted with Program Staff", "Consulted with Parent",
        "Counseling",
    ],
    "Nutrition": [
        "Educational Materials", "Special Diet Menu", "Nutrition Education",
        "Medical Statement for Special Dietary Needs", "WIC",
    ],
    "Transition": [
        "ECI", "Father Involvement", "Fatherhood Conference",
    ],
    "Transportation": [
        "Transportation",
    ],
}

# Optional: map General Service to an internal Department
GENERAL_TO_DEPARTMENT = {
    "Family Services - Emergency/Crisis": "Family Services",
    "Family Well-Being (FPA)": "Family Services",
    "Families as Lifelong Educators (FPA)": "Family Services",
    "Families as Learners (FPA)": "Family Services",
    "Positive Parent-Child Relationship (FPA)": "Family Services",
    "Family Engagement in Transition (FPA)": "Family Services",
    "Family Connections to Peers and Community (FPA)": "Family Services",
    "Families as Advocates and Leaders (FPA)": "Family Services",
    "Education": "Education",
    "Disability": "Disability",
    "Health": "Health",
    "Mental Health": "Mental Health",
    "Nutrition": "Nutrition",
    "Transition": "Education",
    "Transportation": "Transportation",
}

# Normalize “direction” of the action (internal/external/service)
DIRECTION_CANON = {
    "internal": ["internal", "to agency", "in-house", "agency"],
    "external": ["external", "outside", "to outside agency", "community"],
    "service":  ["provided", "service provided", "completed", "delivered"],
}

def canon_direction(val):
    if not isinstance(val, str) or not val.strip():
        return "Unspecified"
    s = val.lower().strip()
    for label, variants in DIRECTION_CANON.items():
        if any(v in s for v in variants):
            return label.capitalize() if label != "service" else "Service"
    if s in ("internal", "external", "service"):
        return s.capitalize()
    return "Unspecified"

# ==============================
# HELPERS
# ==============================
def pick_best_sheet(xls: pd.ExcelFile) -> str:
    prefs = ["service", "activity", "referral", "tracking"]
    for name in xls.sheet_names:
        low = name.lower()
        if any(p in low for p in prefs):
            return name
    return xls.sheet_names[0]

def find_col(columns, *keywords, default=None):
    for c in columns:
        low = str(c).strip().lower()
        if any(k in low for k in keywords):
            return c
    return default

def coerce_to_dt(v):
    if pd.isna(v):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return from_excel(v)
        except Exception:
            return None
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except Exception:
                continue
    return None

def to_general_service(detail: str) -> str:
    if not isinstance(detail, str) or not detail.strip():
        return "Other"
    s = detail.strip().lower()
    for general, details in GENERAL_DETAIL_MAP.items():
        for d in details:
            if d and d.lower() in s:
                return general
    return "Other"

# ==============================
# FILE UPLOAD
# ==============================
uploaded = st.file_uploader("Upload file (e.g., PLite_F_Activity_Service_Tracking.xlsx)", type=["xlsx"])
if not uploaded:
    st.stop()

# Read the most relevant sheet
xls = pd.ExcelFile(uploaded)
sheet_name = pick_best_sheet(xls)
df = pd.read_excel(xls, sheet_name=sheet_name)
df.columns = [str(c).strip() for c in df.columns]

# Key columns (fuzzy)
pid_col       = find_col(df.columns, "pid", "participant id", "child id", default=None)
name_col      = find_col(df.columns, "name", "child", "student", default=None)
detail_col    = find_col(df.columns, "detailed service", "service", "activity/service", "referral type", "service type", "category", default=None)
dept_col      = find_col(df.columns, "department", "dept", default=None)
direction_col = find_col(df.columns, "direction", "internal", "external", "service", "provided", "to agency", "outside", default=None)

# Try to identify a primary date column
date_cols = [c for c in df.columns if re.search(r"date", str(c), re.IGNORECASE)]
primary_date_col = date_cols[0] if date_cols else None

# Coerce dates
if primary_date_col:
    df[primary_date_col] = df[primary_date_col].apply(coerce_to_dt)

# Ensure Detailed Service and General Service columns
if detail_col is None:
    detail_col = "Detailed Service"
    if detail_col not in df.columns:
        df[detail_col] = ""
df["Detailed Service"] = df[detail_col].astype(str).fillna("")
df["General Service"]  = df["Detailed Service"].apply(to_general_service)

# Department resolution
if dept_col and dept_col in df.columns:
    df["Department"] = df[dept_col].fillna("")
    df.loc[df["Department"].str.strip() == "", "Department"] = df["General Service"].map(GENERAL_TO_DEPARTMENT).fillna("Other")
else:
    df["Department"] = df["General Service"].map(GENERAL_TO_DEPARTMENT).fillna("Other")

# Direction normalization
if direction_col and direction_col in df.columns:
    df["Direction"] = df[direction_col].apply(canon_direction)
else:
    df["Direction"] = "Unspecified"

# ==============================
# WRITE FORMATTED WORKBOOK (Sheet1)
# ==============================
title_text = "Referrals / Service Tracking Report 2025–2026"
central_now = datetime.now(ZoneInfo("America/Chicago"))
timestamp_text = central_now.strftime("Generated on %B %d, %Y at %I:%M %p %Z")

temp_path = "Referrals_Cleaned.xlsx"
with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
    pd.DataFrame([[title_text]]).to_excel(writer, index=False, header=False, startrow=0, sheet_name="Referrals")
    pd.DataFrame([[timestamp_text]]).to_excel(writer, index=False, header=False, startrow=1, sheet_name="Referrals")
    df.to_excel(writer, index=False, startrow=3, sheet_name="Referrals")

wb = load_workbook(temp_path)
ws = wb["Referrals"]

filter_row = 4
data_start = filter_row + 1
data_end = ws.max_row
max_col = ws.max_column

# Freeze panes: keep PID + headers visible
ws.freeze_panes = "B5"

# AutoFilter
ws.auto_filter.ref = f"A{filter_row}:{get_column_letter(max_col)}{data_end}"

# Merge & center Title + Timestamp across all columns
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

title_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
ts_fill    = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

tcell = ws.cell(row=1, column=1)
tcell.value = title_text
tcell.font = Font(size=14, bold=True)
tcell.alignment = Alignment(horizontal="center", vertical="center")
tcell.fill = title_fill

scell = ws.cell(row=2, column=1)
scell.value = timestamp_text
scell.font = Font(size=10, italic=True, color="555555")
scell.alignment = Alignment(horizontal="center", vertical="center")
scell.fill = ts_fill

# Header styling: dark blue + white bold + wrapped
header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
for cell in ws[filter_row]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = header_fill

# Borders & date formatting
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
for r in range(data_start, data_end + 1):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        hdr = ws.cell(row=filter_row, column=c).value
        if isinstance(hdr, str) and "date" in hdr.lower():
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "m/d/yy"

# Find Name column index for placing totals after it
name_col_idx = None
for c in range(1, max_col + 1):
    ws.column_dimensions[get_column_letter(c)].width = 14
    hdr = ws.cell(row=filter_row, column=c).value
    if isinstance(hdr, str) and "name" in hdr.lower():
        name_col_idx = c
if name_col_idx is None:
    name_col_idx = 2

# Grand Total at the bottom (no extra wording)
total_row = ws.max_row + 2
ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

center = Alignment(horizontal="center", vertical="center")
for c in range(1, max_col + 1):
    if c <= name_col_idx:
        continue
    count = 0
    for r in range(data_start, data_end + 1):
        val = ws.cell(row=r, column=c).value
        if val not in (None, ""):
            count += 1
    cell = ws.cell(row=total_row, column=c, value=count)
    cell.alignment = center
    cell.font = Font(bold=True)
    cell.border = Border(top=Side(style="thin"))

# ==============================
# SHEET 2: CHARTS
# ==============================
chart_ws = wb.create_sheet("Charts")

# 1) Referrals by Department
dept_counts = df["Department"].fillna("Other").value_counts().sort_index().reset_index()
dept_counts.columns = ["Department", "Count"]

chart_ws.append(["Department", "Count"])
for _, row in dept_counts.iterrows():
    chart_ws.append([row["Department"], int(row["Count"])])

# style header
chart_ws["A1"].font = Font(bold=True, color="FFFFFF")
chart_ws["B1"].font = Font(bold=True, color="FFFFFF")
chart_ws["A1"].fill = header_fill
chart_ws["B1"].fill = header_fill
chart_ws.column_dimensions["A"].width = 30
chart_ws.column_dimensions["B"].width = 12

dept_chart = BarChart()
dept_chart.title = "Referrals by Department"
dept_chart.y_axis.title = "Count"
dept_chart.x_axis.title = "Department"
dept_chart.dataLabels = DataLabelList(showVal=True)

data = Reference(chart_ws, min_col=2, min_row=1, max_row=dept_counts.shape[0] + 1)
cats = Reference(chart_ws, min_col=1, min_row=2, max_row=dept_counts.shape[0] + 1)
dept_chart.add_data(data, titles_from_data=True)
dept_chart.set_categories(cats)
chart_ws.add_chart(dept_chart, "D3")

# 2) Referrals by General Service
start_row = chart_ws.max_row + 3
chart_ws.cell(row=start_row, column=1, value="General Service").font = Font(bold=True, color="FFFFFF")
chart_ws.cell(row=start_row, column=2, value="Count").font = Font(bold=True, color="FFFFFF")
chart_ws.cell(row=start_row, column=1).fill = header_fill
chart_ws.cell(row=start_row, column=2).fill = header_fill

gen_counts = df["General Service"].fillna("Other").value_counts().sort_index().reset_index()
gen_counts.columns = ["General Service", "Count"]

for i, (_idx, row) in enumerate(gen_counts.iterrows(), start=start_row + 1):
    chart_ws.cell(row=i, column=1, value=row["General Service"])
    chart_ws.cell(row=i, column=2, value=int(row["Count"]))

gen_chart = BarChart()
gen_chart.title = "Referrals by General Service"
gen_chart.y_axis.title = "Count"
gen_chart.x_axis.title = "General Service"
gen_chart.dataLabels = DataLabelList(showVal=True)

data2 = Reference(chart_ws, min_col=2, min_row=start_row, max_row=start_row + gen_counts.shape[0])
cats2 = Reference(chart_ws, min_col=1, min_row=start_row + 1, max_row=start_row + gen_counts.shape[0])
gen_chart.add_data(data2, titles_from_data=True)
gen_chart.set_categories(cats2)
chart_ws.add_chart(gen_chart, "D20")

# 3) Stacked: General Service by Direction (Internal / External / Service)
start_row2 = chart_ws.max_row + 3
# header row
chart_ws.cell(row=start_row2, column=1, value="General Service").font = Font(bold=True, color="FFFFFF")
chart_ws.cell(row=start_row2, column=1).fill = header_fill

pivot = pd.pivot_table(df, index="General Service", columns="Direction", aggfunc="size", fill_value=0).reset_index()
pivot_cols = list(pivot.columns)

# write headers
for j, col in enumerate(pivot_cols, start=1):
    cell = chart_ws.cell(row=start_row2, column=j, value=str(col))
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill

# write rows
for i, (_, row) in enumerate(pivot.iterrows(), start=start_row2 + 1):
    for j, col in enumerate(pivot_cols, start=1):
        chart_ws.cell(row=i, column=j, value=int(row[col]) if col != "General Service" else row[col])

stack_chart = BarChart()
stack_chart.type = "col"
stack_chart.grouping = "stacked"
stack_chart.title = "General Service by Direction"
stack_chart.y_axis.title = "Count"
stack_chart.x_axis.title = "General Service"
stack_chart.dataLabels = DataLabelList(showVal=True)

min_col = 2
max_col = 1 + (len(pivot_cols) - 1)
min_row = start_row2
max_row = start_row2 + pivot.shape[0]
stack_data = Reference(chart_ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
stack_cats = Reference(chart_ws, min_col=1, min_row=min_row + 1, max_row=max_row)
stack_chart.add_data(stack_data, titles_from_data=True)
stack_chart.set_categories(stack_cats)
chart_ws.add_chart(stack_chart, "K3")

# ==============================
# SAVE & DOWNLOAD
# ==============================
final_output = "Formatted_Referrals_Report.xlsx"
wb.save(final_output)

with open(final_output, "rb") as f:
    st.download_button("📥 Download Formatted Excel", f, file_name=final_output)
